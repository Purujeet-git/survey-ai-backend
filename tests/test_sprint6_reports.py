"""
Test suite for Sprint 6 — Reports & Spreadsheet Automation:
- Excel Assessment Sheet Generator with Preserved Live Formulas
- Word DOCX Survey Loss Report Generator
- Report Synthesis & Timeline Audit Logging
- Export Endpoints (.xlsx and .docx)
"""

import io
import pytest
import openpyxl
from docx import Document as DocxDocument
from uuid import uuid4

import app.database.models  # Ensures all SQLAlchemy models are registered
from app.reports.services.docx_service import WordReportService
from app.reports.services.excel_service import ExcelAssessmentService


def test_excel_assessment_service_with_formula_preservation():
    service = ExcelAssessmentService()
    claim_data = {
        "claim_number": "CLM-HYUNDAI-001",
        "insured_name": "RAMSATI DEVI",
        "registration_number": "JH01EX7415",
        "vehicle_model": "HYUNDAI CRETA",
        "policy_number": "POL-99482103",
    }
    parts = [
        {"part_code": "86551K6000", "description": "BRACKET-FR BUMPER SIDE LH", "qty": 1, "rate": 111.02, "tax": 0.18, "assessed": 111.02, "depr": 0.50},
        {"part_code": "86511K6000", "description": "COVER-FR BUMPER", "qty": 1, "rate": 1483.90, "tax": 0.18, "assessed": 1483.90, "depr": 0.50},
    ]
    labor = [
        {"code": "A10AARER27LNA", "description": "Rear Door denting/repair RH", "tax": 0.18, "claimed": 500.0, "assessed": 500.0},
    ]

    excel_bytes = service.generate_assessment_excel(
        claim_data=claim_data,
        parts_list=parts,
        labor_list=labor,
        less_excess=1000.0,
        salvage_value=500.0,
    )

    assert len(excel_bytes) > 0

    # Load generated excel workbook and verify formulas exist
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
    ws = wb["Survey Assessment"]

    # Verify header metadata
    assert ws["B3"].value == "CLM-HYUNDAI-001"
    assert ws["G3"].value == "RAMSATI DEVI"

    # Verify preserved dynamic Excel formulas
    formula_found = False
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula_found = True
                break

    assert formula_found is True, "Excel spreadsheet must contain preserved dynamic formulas (=SUM, =E*F)."


def test_word_report_service_docx_generation():
    service = WordReportService()
    claim_data = {
        "claim_number": "CLM-HYUNDAI-001",
        "insured_name": "RAMSATI DEVI",
        "registration_number": "JH01EX7415",
        "vehicle_model": "HYUNDAI CRETA",
        "policy_number": "POL-99482103",
        "accident_narrative": "Vehicle collided with barrier resulting in frontal bumper damage.",
    }
    parts = [
        {"part_code": "86551K6000", "description": "BRACKET-FR BUMPER SIDE LH", "claimed": 111.02, "assessed": 111.02},
    ]
    labor = [
        {"code": "A10AARER27LNA", "description": "Rear Door denting/repair RH", "claimed": 500.0, "assessed": 500.0},
    ]
    findings = [
        {"title": "Unsupported Repair", "severity": "HIGH", "description": "Rear door replacement not supported by frontal crash."}
    ]

    docx_bytes = service.generate_survey_report_docx(
        claim_data=claim_data,
        parts_list=parts,
        labor_list=labor,
        findings=findings,
        less_excess=1000.0,
        salvage_value=500.0,
    )

    assert len(docx_bytes) > 0

    # Load Word document and verify content
    doc = DocxDocument(io.BytesIO(docx_bytes))
    full_text = "\n".join([p.text for p in doc.paragraphs])

    assert "MOTOR INSURANCE SURVEYOR LOSS ASSESSMENT REPORT" in full_text
    assert "CLM-HYUNDAI-001" in full_text
    assert "Unsupported Repair" in full_text
