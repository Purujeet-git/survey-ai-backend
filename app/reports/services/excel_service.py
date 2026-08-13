"""
SurveyAI Backend

Module:
Excel Assessment Sheet Generator (with Native Formula Preservation)

Purpose:
Generates surveyor assessment Excel spreadsheets using openpyxl, preserving native dynamic formulas for surveyor review.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelAssessmentService:
    """
    Service for building surveyor Excel assessment calculation sheets with preserved live formulas.
    """

    def generate_assessment_excel(
        self,
        claim_data: dict,
        parts_list: list[dict],
        labor_list: list[dict],
        less_excess: float = 1000.0,
        salvage_value: float = 500.0,
    ) -> bytes:
        """
        Generate .xlsx assessment workbook bytes with preserved formulas.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Survey Assessment"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)

        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # Title Block
        ws.merge_cells("A1:K1")
        ws["A1"] = "MOTOR INSURANCE SURVEYOR ASSESSMENT SHEET"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Claim Metadata Block
        meta_rows = [
            ("Claim Number:", claim_data.get("claim_number", "CLM-9901"), "Insured Name:", claim_data.get("insured_name", "RAMSATI DEVI")),
            ("Policy Number:", claim_data.get("policy_number", "POL-99482103"), "Registration No:", claim_data.get("registration_number", "JH01EX7415")),
            ("Vehicle Make/Model:", claim_data.get("vehicle_model", "HYUNDAI CRETA"), "Date of Loss:", claim_data.get("incident_date", "2026-06-09")),
            ("Workshop / Garage:", claim_data.get("workshop", "RAMA AUTO DEALERS PVT. LTD."), "Surveyor Name:", claim_data.get("surveyor_name", "Official Insurance Surveyor")),
        ]

        row = 3
        for label1, val1, label2, val2 in meta_rows:
            ws[f"A{row}"] = label1
            ws[f"A{row}"].font = bold_font
            ws[f"B{row}"] = val1
            ws[f"B{row}"].font = regular_font

            ws[f"F{row}"] = label2
            ws[f"F{row}"].font = bold_font
            ws[f"G{row}"] = val2
            ws[f"G{row}"].font = regular_font
            row += 1

        row += 1

        # --- Section 1: Parts Assessment Table ---
        ws.merge_cells(f"A{row}:K{row}")
        ws[f"A{row}"] = "PART INVOICE ASSESSMENT"
        ws[f"A{row}"].font = bold_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        part_headers = [
            "S.No", "Part Code", "Description", "HSN/SAC", "Qty", "Rate",
            "Claimed Amt", "Tax %", "Assessed Amt", "Depr %", "Net Amt"
        ]
        
        for col_num, h_text in enumerate(part_headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = h_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        parts_start_row = row + 1
        row += 1

        # Default fallback parts if none supplied
        items = parts_list or [
            {"part_code": "86551K6000", "description": "BRACKET-FR BUMPER SIDE LH", "hsn": "87089900", "qty": 1, "rate": 111.02, "tax": 0.18, "assessed": 111.02, "depr": 0.50},
            {"part_code": "86511K6000", "description": "COVER-FR BUMPER", "hsn": "87089900", "qty": 1, "rate": 1483.90, "tax": 0.18, "assessed": 1483.90, "depr": 0.50},
            {"part_code": "86300K6010", "description": "EMBLEM-SYMBOL MARK", "hsn": "87089900", "qty": 1, "rate": 613.56, "tax": 0.18, "assessed": 613.56, "depr": 0.00},
            {"part_code": "83404C4010", "description": "REGULATOR ASSY-RR DR WDO RH", "hsn": "87089900", "qty": 1, "rate": 753.38, "tax": 0.18, "assessed": 753.38, "depr": 0.00},
        ]

        for i, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=i).font = regular_font
            ws.cell(row=row, column=2, value=item.get("part_code", "")).font = regular_font
            ws.cell(row=row, column=3, value=item.get("description", "")).font = regular_font
            ws.cell(row=row, column=4, value=item.get("hsn", "87089900")).font = regular_font
            ws.cell(row=row, column=5, value=item.get("qty", 1)).font = regular_font
            ws.cell(row=row, column=6, value=float(item.get("rate", 0.0))).font = regular_font
            
            # Preserved Formula for Claimed Amt: Qty * Rate
            ws.cell(row=row, column=7, value=f"=E{row}*F{row}").font = regular_font
            
            ws.cell(row=row, column=8, value=float(item.get("tax", 0.18))).font = regular_font
            ws.cell(row=row, column=9, value=float(item.get("assessed", item.get("rate", 0.0)))).font = regular_font
            ws.cell(row=row, column=10, value=float(item.get("depr", 0.0))).font = regular_font
            
            # Preserved Formula for Net Amt: Assessed * (1 - Depr%)
            ws.cell(row=row, column=11, value=f"=I{row}*(1-J{row})").font = bold_font
            
            for c in range(1, 12):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        parts_end_row = row - 1

        # Parts Subtotal Row with Excel Formula
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = "Sub Total Parts (Net Assessed):"
        ws[f"A{row}"].font = bold_font
        ws[f"A{row}"].alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=11, value=f"=SUM(K{parts_start_row}:K{parts_end_row})").font = bold_font
        ws.cell(row=row, column=11).fill = total_fill
        parts_subtotal_cell = f"K{row}"
        row += 2

        # --- Section 2: Labour & Services Table ---
        ws.merge_cells(f"A{row}:K{row}")
        ws[f"A{row}"] = "LABOUR & SERVICES ASSESSMENT"
        ws[f"A{row}"].font = bold_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        labor_headers = ["S.No", "Labor Code", "Description", "HSN/SAC", "Tax %", "Claimed Amt", "Assessed Amt"]
        for col_num, h_text in enumerate(labor_headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = h_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        labor_start_row = row + 1
        row += 1

        l_items = labor_list or [
            {"code": "A10AARER27LNA", "description": "Rear Door denting/repair RH", "hsn": "998729", "tax": 0.18, "claimed": 500.0, "assessed": 500.0},
            {"code": "A10AARFBPF07R", "description": "COVER- FR BUMPER R & R", "hsn": "998729", "tax": 0.18, "claimed": 451.0, "assessed": 451.0},
            {"code": "A10AAWBBPF13FH", "description": "FR Bumper (Water Borne Paint)", "hsn": "998729", "tax": 0.18, "claimed": 5687.0, "assessed": 5687.0},
        ]

        for i, l_item in enumerate(l_items, 1):
            ws.cell(row=row, column=1, value=i).font = regular_font
            ws.cell(row=row, column=2, value=l_item.get("code", "")).font = regular_font
            ws.cell(row=row, column=3, value=l_item.get("description", "")).font = regular_font
            ws.cell(row=row, column=4, value=l_item.get("hsn", "998729")).font = regular_font
            ws.cell(row=row, column=5, value=float(l_item.get("tax", 0.18))).font = regular_font
            ws.cell(row=row, column=6, value=float(l_item.get("claimed", 0.0))).font = regular_font
            ws.cell(row=row, column=7, value=float(l_item.get("assessed", 0.0))).font = bold_font
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

        labor_end_row = row - 1

        # Labor Subtotal Row with Excel Formula
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "Sub Total Labor (Assessed):"
        ws[f"A{row}"].font = bold_font
        ws[f"A{row}"].alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=7, value=f"=SUM(G{labor_start_row}:G{labor_end_row})").font = bold_font
        ws.cell(row=row, column=7).fill = total_fill
        labor_subtotal_cell = f"G{row}"
        row += 2

        # --- Section 3: Final Assessment Summary with Preserved Excel Formulas ---
        ws.merge_cells(f"A{row}:K{row}")
        ws[f"A{row}"] = "FINAL SURVEY ASSESSMENT SUMMARY"
        ws[f"A{row}"].font = bold_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        summary_rows = [
            ("Total Assessed Parts:", f"={parts_subtotal_cell}"),
            ("Total Assessed Labor:", f"={labor_subtotal_cell}"),
            ("Gross Total Assessment:", f"=K{row}+K{row+1}"),
            ("Add: GST @ 18%:", f"=K{row+2}*0.18"),
            ("Total Assessment with Tax:", f"=K{row+2}+K{row+3}"),
            ("Less: Salvage Value:", f"={salvage_value}"),
            ("Less: Compulsory Excess:", f"={less_excess}"),
            ("NET PAYABLE AMOUNT:", f"=K{row+4}-K{row+5}-K{row+6}"),
        ]

        net_payable_row = row + len(summary_rows) - 1

        for idx, (label_text, formula_expr) in enumerate(summary_rows):
            ws.merge_cells(f"F{row}:J{row}")
            ws[f"F{row}"] = label_text
            ws[f"F{row}"].font = bold_font
            ws[f"F{row}"].alignment = Alignment(horizontal="right")
            
            cell = ws.cell(row=row, column=11, value=formula_expr)
            cell.font = bold_font
            cell.border = thin_border
            if idx == len(summary_rows) - 1:
                cell.fill = total_fill
                cell.font = Font(name="Calibri", size=11, bold=True, color="1F497D")
            row += 1

        # Adjust column widths
        from openpyxl.utils import get_column_letter
        col_widths = {1: 8, 2: 18, 3: 35, 4: 12, 5: 8, 6: 12, 7: 14, 8: 10, 9: 14, 10: 10, 11: 16}
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width


        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
